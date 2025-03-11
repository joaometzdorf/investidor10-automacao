from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from time import sleep
import openpyxl


def fechar_anuncio(driver):
    try:
        botao_fechar_anuncio = driver.find_element(
            By.XPATH,
            "//div[contains(@class, 'show')]//div[contains(@class, 'modal-default-container')]//button[contains(@class, 'modal-close')]",
        )  # Botão de fechar anúncio

        botao_fechar_anuncio.click()
        sleep(2)
        print("Anúncio fechado.")
        return
    except NoSuchElementException:
        print("Não há anúncio para fechar.")
        sleep(1)


# Carregar o arquivo Excel
arquivo_dados_fiis = openpyxl.load_workbook("dados_fiis.xlsx")
planilha_dados_fiis = arquivo_dados_fiis["FIIs"]

# Encontrar a primeira linha vazia (começando na linha 2)
linha_atual = 2
while planilha_dados_fiis.cell(row=linha_atual, column=1).value:
    linha_atual += 1

# Iniciar o driver
driver = webdriver.Chrome()
driver.maximize_window()

# Acessar o site
driver.get("https://investidor10.com.br")
sleep(2)

# Navegar até a seção de FIIs
botao_fiis = driver.find_elements(By.XPATH, "//li[@class='has-children']")[1]
botao_fiis.click()
sleep(2)

botao_abrir_todos = driver.find_element(
    By.XPATH, "//a[@href='https://investidor10.com.br/fiis/']"
)
botao_abrir_todos.click()
sleep(2)

# Loop para percorrer todas as páginas
while True:
    # Pegar os links das FIIs antes de abri-las
    links_fiis = [
        card.get_attribute("href")
        for card in driver.find_elements(By.XPATH, "//div[@class='actions fii']//a")
    ]

    # Abrir cada FII em uma nova aba
    for link in links_fiis:
        driver.execute_script(f"window.open('{link}');")  # Abrir nova aba
        driver.switch_to.window(driver.window_handles[-1])  # Mudar para a nova aba
        sleep(3)

        try:
            # Coletar os dados
            nome_fii = driver.find_element(By.XPATH, "//h1").text
            cota_fii = driver.find_element(
                By.XPATH, "//div[@class='_card cotacao']//span[@class='value']"
            ).text
            tipo_de_fundo_fii = driver.find_elements(
                By.XPATH, "//div[@class='desc']//div[@class='value']//span"
            )[5].text
            valor_patrimonial_fii = driver.find_elements(
                By.XPATH, "//div[@class='desc']//div[@class='value']//span"
            )[13].text
            liquidez_diaria_fii = driver.find_element(
                By.XPATH, "//div[@class='_card val']//div[@class='_card-body']//span"
            ).text
            pvp_fii = driver.find_element(
                By.XPATH, "//div[@class='_card vp']//div[@class='_card-body']//span"
            ).text
            dividend_yield_fii = driver.find_element(
                By.XPATH, "//div[@class='_card dy']//div[@class='_card-body']//span"
            ).text
            vacancia_fii = driver.find_elements(
                By.XPATH, "//div[@class='desc']//div[@class='value']//span"
            )[9].text

            # Adicionar os dados na planilha
            planilha_dados_fiis.append(
                [
                    nome_fii,
                    cota_fii,
                    tipo_de_fundo_fii,
                    valor_patrimonial_fii,
                    liquidez_diaria_fii,
                    pvp_fii,
                    dividend_yield_fii,
                    vacancia_fii,
                ]
            )
            planilha_dados_fiis.cell(row=linha_atual, column=1).value = nome_fii
            planilha_dados_fiis.cell(row=linha_atual, column=2).value = cota_fii
            planilha_dados_fiis.cell(row=linha_atual, column=3).value = (
                tipo_de_fundo_fii
            )
            planilha_dados_fiis.cell(row=linha_atual, column=4).value = (
                valor_patrimonial_fii
            )
            planilha_dados_fiis.cell(row=linha_atual, column=5).value = (
                liquidez_diaria_fii
            )
            planilha_dados_fiis.cell(row=linha_atual, column=6).value = pvp_fii
            planilha_dados_fiis.cell(row=linha_atual, column=7).value = (
                dividend_yield_fii
            )
            planilha_dados_fiis.cell(row=linha_atual, column=8).value = vacancia_fii
            linha_atual += 1  # Avançar para a próxima linha

        except Exception as e:
            print(f"Erro ao coletar dados: {e}")

        # Fechar a aba da FII e voltar para a principal
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        sleep(3)

    # Fechar anúncio se houver
    fechar_anuncio(driver)
    sleep(2)
    driver.execute_script("window.scrollBy(0, 2500);")  # Descer a página
    sleep(3)

    # Tentar ir para a próxima página
    try:
        botao_proxima_pagina = driver.find_element(
            By.XPATH, "//li[@class='pagination-item next']//a[@class='pagination-link']"
        )

        pagina_ativa = driver.find_element(
            By.XPATH,
            "//li[@class='pagination-item active']//a[@class='pagination-link']",
        ).text

        if pagina_ativa != "13":
            botao_proxima_pagina.click()
            sleep(3)
        else:
            break
    except Exception as e:
        print(f"Não há mais páginas. Erro: {e}")
        break

# Salvar os dados no Excel
arquivo_dados_fiis.save("dados_fiis.xlsx")
arquivo_dados_fiis.close()
driver.quit()  # Fechar o navegador
